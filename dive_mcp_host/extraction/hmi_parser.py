"""TIA Portal HMI screen data extractor.

Parses RDF binary screen files from a TIA Portal project directory,
extracts UI elements, JavaScript events, PLC tag bindings, and screen
navigation maps.  Returns structured results as Pydantic models from
``dive_mcp_host.extraction.models``.

Pure parsing functions (RDF binary regex, tag extraction, element
classification) are kept side-effect-free.  The sole entry point is
:func:`parse_hmi_project`.
"""

from __future__ import annotations

import logging
import os
import re
from collections import defaultdict
from pathlib import Path

from dive_mcp_host.extraction.models import (
    ElementEvent,
    HmiExtraction,
    HmiSummary,
    ScreenElement,
    ScreenResult,
    TagBinding,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# HMITags.xlsx reader (lazy openpyxl import)
# ---------------------------------------------------------------------------

def load_hmi_tags(xlsx_path: str | Path) -> dict[str, dict]:
    """Load *HMITags.xlsx* and return a dict keyed by HMI tag name.

    Each value is a detail dict with keys like ``plc_tag``,
    ``data_type``, ``connection``, etc.

    Returns an empty dict when openpyxl is missing or the file does not
    exist.
    """
    try:
        import openpyxl  # noqa: WPS433  – lazy import
    except ImportError:
        logger.warning("openpyxl not installed; HMI tag resolution disabled")
        return {}

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        logger.warning("HMITags.xlsx not found at %s", xlsx_path)
        return {}

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    tags: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h else "" for h in rows[0]]

        for row in rows[1:]:
            values = list(row)
            row_dict: dict[str, object] = {}
            for idx, h in enumerate(headers):
                if idx < len(values):
                    row_dict[h] = values[idx] if values[idx] is not None else ""

            name = row_dict.get("Name", "")
            if not name:
                continue

            tags[str(name)] = {
                "hmi_tag": str(name),
                "path": str(row_dict.get("Path", "")),
                "connection": str(row_dict.get("Connection", "")),
                "plc_tag": str(row_dict.get("PLC tag", "")),
                "data_type": str(row_dict.get("DataType", "")),
                "hmi_data_type": str(row_dict.get("HMI DataType", "")),
                "length": row_dict.get("Length", ""),
                "access_method": str(row_dict.get("Access Method", "")),
                "address": str(row_dict.get("Address", "")),
                "start_value": str(row_dict.get("Start value", "")),
                "comment": str(row_dict.get("Comment [en-US]", "")),
                "acquisition_mode": str(row_dict.get("Acquisition mode", "")),
                "acquisition_cycle": str(row_dict.get("Acquisition cycle", "")),
                "limit_upper": str(row_dict.get("Limit Upper 2", "")),
                "limit_lower": str(row_dict.get("Limit Lower 2", "")),
            }

    wb.close()
    logger.info("Loaded %d HMI tags from %s", len(tags), xlsx_path.name)
    return tags


# ---------------------------------------------------------------------------
# RDF binary parser helpers
# ---------------------------------------------------------------------------

def extract_js_functions(data: bytes) -> list[dict[str, str]]:
    """Extract JavaScript ``export async function`` blocks from RDF bytes."""
    js_pattern = rb"export\s+async\s+function\s+(\w+)\s*\([^)]*\)\s*\{"
    functions: list[dict[str, str]] = []

    for match in re.finditer(js_pattern, data):
        func_name = match.group(1).decode("utf-8", errors="ignore")
        brace_count = 0
        func_start = data.index(b"{", match.start())
        pos = func_start
        while pos < len(data) and pos - func_start < 10000:
            if data[pos : pos + 1] == b"{":
                brace_count += 1
            elif data[pos : pos + 1] == b"}":
                brace_count -= 1
                if brace_count == 0:
                    func_body = data[match.start() : pos + 1].decode(
                        "utf-8", errors="ignore"
                    )
                    functions.append({"name": func_name, "body": func_body})
                    break
            pos += 1
    return functions


def extract_plc_tags(js_code: str) -> list[str]:
    """Return sorted unique PLC tag references found in *js_code*."""
    tags: set[str] = set()
    patterns = [
        r'GetTagValue\("([^"]+)"\)',
        r'SetTagValue\("([^"]+)"',
        r'SetBitInTag\("([^"]+)"',
        r'ResetBitInTag\("([^"]+)"',
        r'TagValue\("([^"]+)"\)',
        r'"([A-Z][A-Z0-9_]*_[A-Z0-9_ .\-]+)"',
        r'"(DB_[A-Z_0-9][A-Z_0-9 .\-]+)"',
    ]
    for pattern in patterns:
        for m in re.finditer(pattern, js_code):
            tags.add(m.group(1).strip())
    return sorted(tags)


def extract_screen_navigation(js_code: str) -> list[str]:
    """Extract ``ChangeScreen("...")`` targets from JS code."""
    return [m.group(1) for m in re.finditer(r'ChangeScreen\("([^"]+)"', js_code)]


def extract_element_details(data: bytes) -> list[dict]:
    """Extract UI element names and types from RDF binary data."""
    elements: list[dict] = []
    elem_pattern = (
        rb"((?:Button|Text box|Screen window|IO field|Bar|Symbolic IO field"
        rb"|Status display|Switch|Slider|Clock|Date|Text list|Image"
        rb"|Rectangle|Line|Circle|Faceplate)_(?:\w+))"
    )

    event_suffixes = (
        "_OnUp", "_OnDown", "_OnClick", "_OnChange", "_OnLoaded",
        "_OnMouseEnter", "_OnMouseLeave", "_OnFocus", "_OnBlur",
        "_OnKeyPress", "_OnKeyDown", "_OnKeyUp", "_OnShow", "_OnHide",
    )

    seen_names: set[str] = set()

    type_map = {
        "Button": "button",
        "Text box": "text_display",
        "Screen window": "screen_window",
        "IO field": "io_field",
        "Bar": "bar_graph",
        "Symbolic IO field": "symbolic_io_field",
        "Status display": "status_display",
        "Switch": "switch",
        "Slider": "slider",
        "Rectangle": "rectangle",
        "Line": "line",
        "Circle": "circle",
        "Image": "image",
        "Faceplate": "faceplate",
    }

    for match in re.finditer(elem_pattern, data):
        elem_name = match.group(1).decode("utf-8", errors="ignore")
        elem_start = match.start()

        # Skip JS function names (e.g. Button_9_OnUp)
        if any(elem_name.endswith(s) for s in event_suffixes):
            continue
        if elem_name in seen_names:
            continue
        seen_names.add(elem_name)

        elem_type = "unknown"
        for prefix, t in type_map.items():
            if elem_name.startswith(prefix):
                elem_type = t
                break

        # Scan the binary region after the element name for tag references
        region = data[elem_start : min(elem_start + 200, len(data))]
        tag_refs: list[str] = []
        region_str = region.decode("utf-8", errors="ignore")
        for tm in re.finditer(r"([A-Z][A-Z0-9_]*_[A-Z0-9_ .\-]{4,})", region_str):
            tag_refs.append(tm.group(1).strip())

        elements.append(
            {
                "name": elem_name,
                "type": elem_type,
                "events": [],
                "tag_references_in_region": tag_refs,
            }
        )

    return elements


# ---------------------------------------------------------------------------
# Screen-level orchestrator
# ---------------------------------------------------------------------------

def parse_screen_rdf(
    filepath: str | Path,
    hmi_tags: dict[str, dict],
) -> ScreenResult | None:
    """Parse a single screen RDF file and return a :class:`ScreenResult`.

    Returns ``None`` when the file is too small to contain useful data.
    """
    filepath = Path(filepath)
    with open(filepath, "rb") as fh:
        data = fh.read()

    if len(data) < 10:
        return None

    # -- Screen name --------------------------------------------------------
    all_strings: list[str] = []
    for m in re.finditer(rb"[\x20-\x7e]{3,}", data):
        s = m.group().decode("ascii", errors="ignore").strip()
        if s:
            all_strings.append(s)

    screen_name = _detect_screen_name(all_strings)

    # -- Elements -----------------------------------------------------------
    elements = extract_element_details(data)

    # -- JavaScript functions -----------------------------------------------
    js_functions = extract_js_functions(data)
    all_js_code = "\n".join(f["body"] for f in js_functions)
    plc_tags_from_js = extract_plc_tags(all_js_code)
    navigations = extract_screen_navigation(all_js_code)

    # -- Link JS functions to elements --------------------------------------
    for func in js_functions:
        func_name = func["name"]
        func_tags = extract_plc_tags(func["body"])
        func_navs = extract_screen_navigation(func["body"])

        for elem in elements:
            elem_key = elem["name"].replace(" ", "_")
            if not func_name.startswith(elem_key + "_"):
                continue

            event_type = _event_type_from_suffix(func_name)

            elem["events"].append(
                {
                    "function": func_name,
                    "event_type": event_type,
                    "code": func["body"],
                    "plc_tags": func_tags,
                    "resolved_plc_tags": _resolve_tags(func_tags, hmi_tags),
                    "navigates_to": func_navs,
                }
            )

    # -- OnLoaded screen event ----------------------------------------------
    on_loaded: str | None = None
    for func in js_functions:
        if "OnLoaded" in func["name"]:
            on_loaded = func["body"]

    # -- Enrich elements with HMITag data -----------------------------------
    for elem in elements:
        all_elem_tags: set[str] = set()
        for ev in elem.get("events", []):
            for t in ev.get("plc_tags", []):
                all_elem_tags.add(t)
        for t in elem.get("tag_references_in_region", []):
            all_elem_tags.add(t)

        elem["hmi_tags"] = _resolve_tags(sorted(all_elem_tags), hmi_tags)
        elem["io_role"] = classify_io_role(elem)

    # -- Build ScreenResult model -------------------------------------------
    screen_elements = _build_screen_elements(elements)

    return ScreenResult(
        screen_name=screen_name or "UNKNOWN",
        file=filepath.name,
        element_count=len(elements),
        elements=screen_elements,
        element_summary=summarize_elements(elements),
        javascript_functions=[
            {"name": f["name"], "body": f["body"]} for f in js_functions
        ],
        screen_navigations=navigations,
        on_loaded_event=on_loaded,
        plc_tags_referenced=plc_tags_from_js,
    )


# ---------------------------------------------------------------------------
# Element classification
# ---------------------------------------------------------------------------

def classify_io_role(elem: dict) -> str:
    """Classify an element as input, output, display, navigation, or control."""
    events = elem.get("events", [])
    elem_type = elem.get("type", "")

    if elem_type == "button":
        for ev in events:
            if ev.get("navigates_to"):
                return "navigation_button"
            if ev.get("plc_tags"):
                return "control_button (writes to PLC)"
        return "button"

    if elem_type == "io_field":
        has_write = False
        for ev in events:
            code = ev.get("code", "")
            if "Set" in code or "Write" in code:
                has_write = True
        if has_write:
            return "input (writes to PLC)"
        return "output (reads from PLC)"

    if elem_type == "symbolic_io_field":
        return "selection_input"

    if elem_type == "screen_window":
        return "screen_container"

    if elem_type in ("text_display", "rectangle", "line", "circle", "image"):
        return "display/static"

    if elem_type == "bar_graph":
        return "output (visualizes PLC value)"

    if elem_type == "status_display":
        return "output (PLC status)"

    if elem_type == "switch":
        return "input (writes to PLC)"

    return "display/static"


def summarize_elements(elements: list[dict]) -> dict[str, int]:
    """Create a summary dict of element type counts on a screen."""
    summary: dict[str, int] = defaultdict(int)
    for e in elements:
        summary[e.get("type", "unknown")] += 1
        if e.get("events"):
            summary["with_events"] += 1
        if e.get("hmi_tags"):
            summary["with_tag_bindings"] += 1
        io = e.get("io_role", "")
        if "input" in io:
            summary["inputs"] += 1
        elif "output" in io or "display" in io:
            summary["outputs"] += 1
        elif "control" in io or "navigation" in io:
            summary["controls"] += 1
    return dict(summary)


# ---------------------------------------------------------------------------
# HMI instance discovery
# ---------------------------------------------------------------------------

def list_hmi_instances(project_path: str | Path) -> dict[str, dict]:
    """Return all HMI instances found under ``IM/HMI/I/``.

    Each key is the instance directory name (numeric string); each value
    contains device metadata and screen count.
    """
    project_path = Path(project_path)
    base_dir = project_path / "IM" / "HMI" / "I"
    if not base_dir.exists():
        return {}

    instances: dict[str, dict] = {}
    for entry in sorted(base_dir.iterdir()):
        if not entry.is_dir():
            continue
        for sub in ("Saved", "Context"):
            dt_path = entry / sub / "DownloadTask.xml"
            if not dt_path.exists():
                continue
            try:
                content = dt_path.read_text(encoding="utf-8", errors="ignore")
                info: dict[str, str] = {}
                for field in (
                    "ESDeviceName",
                    "RtProjectFolderName",
                    "DeviceType",
                    "DeviceVersion",
                    "CreationTime",
                ):
                    m = re.search(rf'{field}="([^"]*)"', content)
                    if m:
                        info[field] = m.group(1)

                screens_dir = entry / sub / "screens"
                screen_count = 0
                if screens_dir.exists():
                    screen_count = sum(
                        1
                        for f in screens_dir.iterdir()
                        if f.name.startswith("screen_") and f.suffix == ".rdf"
                    )

                instances[entry.name] = {
                    "source": sub,
                    "device_type": info.get("DeviceType", "Unknown"),
                    "device_version": info.get("DeviceVersion", "?"),
                    "device_name": info.get("RtProjectFolderName", "Unknown"),
                    "es_device": info.get("ESDeviceName", "?"),
                    "creation_time": info.get("CreationTime", "?"),
                    "screen_count": screen_count,
                }
            except Exception:
                instances[entry.name] = {"source": sub, "error": "could not parse"}
            break  # only need the first hit (Saved preferred over Context)

    return instances


def find_screen_files(
    project_path: str | Path,
    instance_id: int | None = None,
) -> list[Path]:
    """Discover screen ``*.rdf`` files under a project's HMI instance dirs.

    When *instance_id* is ``None`` the newest instance (by
    ``creation_time``) is selected automatically.
    """
    project_path = Path(project_path)
    base_dir = project_path / "IM" / "HMI" / "I"
    if not base_dir.exists():
        return []

    if instance_id is not None:
        instance_dirs = [base_dir / str(instance_id)]
    else:
        instances = list_hmi_instances(project_path)
        if instances:
            latest = max(
                (
                    (k, v)
                    for k, v in instances.items()
                    if "creation_time" in v
                ),
                key=lambda kv: kv[1].get("creation_time", ""),
                default=(None, {}),
            )
            if latest[0] is not None:
                instance_dirs = [base_dir / latest[0]]
                logger.info(
                    "Auto-selected instance %s (newest: %s)",
                    latest[0],
                    latest[1].get("creation_time", "?"),
                )
            else:
                instance_dirs = [base_dir]
        else:
            instance_dirs = [base_dir]

    screen_files: set[Path] = set()
    for search_dir in instance_dirs:
        if not search_dir.exists():
            continue
        for root, _dirs, files in os.walk(search_dir):
            for fname in files:
                if fname.startswith("screen_") and fname.endswith(".rdf"):
                    full_path = Path(root) / fname
                    # Only collect files under a Saved/ tree
                    if "Saved" in full_path.parts:
                        screen_files.add(full_path)

    return sorted(screen_files)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_hmi_project(
    project_path: str | Path,
    hmi_tags_path: str | Path | None = None,
    instance_id: int | None = None,
) -> HmiExtraction:
    """Parse all HMI screens in a TIA Portal project.

    Parameters
    ----------
    project_path:
        Root directory of the TIA Portal project (containing ``IM/HMI/``).
    hmi_tags_path:
        Optional path to ``HMITags.xlsx``.  When provided, tag bindings
        are resolved to PLC tag details.
    instance_id:
        Specific HMI instance ID.  When ``None`` the newest instance is
        selected automatically.

    Returns
    -------
    HmiExtraction
        Structured extraction result ready for serialization or API
        response.
    """
    project_path = Path(project_path)

    # -- Load HMI tags (optional) -------------------------------------------
    hmi_tags: dict[str, dict] = {}
    if hmi_tags_path is not None:
        hmi_tags = load_hmi_tags(hmi_tags_path)

    # -- Discover screen RDF files ------------------------------------------
    screen_files = find_screen_files(project_path, instance_id=instance_id)
    logger.info("Found %d screen RDF files", len(screen_files))

    # -- Parse each screen (fault-tolerant) ---------------------------------
    screens: list[ScreenResult] = []
    errors: list[dict] = []

    for sf in screen_files:
        try:
            result = parse_screen_rdf(sf, hmi_tags)
            if result is not None:
                screens.append(result)
                logger.debug(
                    "Parsed screen %s (%d elements)",
                    result.screen_name,
                    result.element_count,
                )
        except Exception as exc:
            errors.append({"file": str(sf), "error": str(exc)})
            logger.warning("Failed to parse screen %s: %s", sf, exc)

    # -- Build navigation map -----------------------------------------------
    navigation_map: dict[str, list[str]] = {}
    for scr in screens:
        if scr.screen_navigations:
            navigation_map[scr.screen_name] = scr.screen_navigations

    # -- Build PLC tag index ------------------------------------------------
    raw_tag_index: dict[str, list[str]] = defaultdict(list)
    for scr in screens:
        for tag in scr.plc_tags_referenced:
            raw_tag_index[tag].append(scr.screen_name)

    plc_tag_index: dict[str, dict] = {}
    for tag_name, used_in in sorted(raw_tag_index.items()):
        tag_detail = hmi_tags.get(tag_name)
        plc_tag_index[tag_name] = {
            "plc_tag": tag_detail["plc_tag"] if tag_detail else "(not in tag table)",
            "data_type": tag_detail["data_type"] if tag_detail else "?",
            "connection": tag_detail["connection"] if tag_detail else "?",
            "used_in_screens": used_in,
        }

    # -- Device info --------------------------------------------------------
    hmi_device: dict | None = None
    if screen_files:
        first = str(screen_files[0]).replace("\\", "/")
        m = re.search(r"/I/(\d+)/", first)
        if m:
            instances = list_hmi_instances(project_path)
            hmi_device = dict(instances.get(m.group(1), {}))
            hmi_device["instance_id"] = m.group(1)

    # -- Summary ------------------------------------------------------------
    total_elements = sum(s.element_count for s in screens)
    total_events = sum(
        1 for s in screens for e in s.elements if e.events
    )
    total_tag_bindings = sum(
        len(e.tag_bindings) for s in screens for e in s.elements
    )
    total_js_functions = sum(len(s.javascript_functions) for s in screens)

    summary = HmiSummary(
        total_screens=len(screens),
        total_elements=total_elements,
        total_elements_with_events=total_events,
        total_tag_bindings=total_tag_bindings,
        total_js_functions=total_js_functions,
        total_unique_plc_tags=len(plc_tag_index),
        total_navigation_links=sum(len(v) for v in navigation_map.values()),
    )

    logger.info(
        "HMI extraction complete: %d screens, %d elements, %d PLC tags",
        summary.total_screens,
        summary.total_elements,
        summary.total_unique_plc_tags,
    )

    return HmiExtraction(
        source_project=project_path.name,
        summary=summary,
        screens=screens,
        navigation_map=navigation_map,
        plc_tag_index=plc_tag_index,
        hmi_device=hmi_device,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

_SKIP_PATTERNS = (
    "LAYER", "MODULE", "FONT", "COLOR", "IMAGE", "STYLE",
    "TOOLBAR", "BUTTON_", "FACEPLATE", "TEMPLATE", "VERSION",
    "AUTHOR", "CREATED", "MODIFIED", "SIEMENS",
)


def _detect_screen_name(all_strings: list[str]) -> str | None:
    """Heuristic: pick the most plausible screen name from RDF strings."""
    for s in all_strings:
        if any(skip in s.upper() for skip in _SKIP_PATTERNS):
            continue
        if (
            re.match(r"^[A-Z0-9][A-Z0-9_ ]{2,}$", s)
            and 3 <= len(s) < 60
        ):
            return s

    # Broader fallback (mixed-case names, still skipping noise)
    for s in all_strings:
        if any(skip in s.upper() for skip in _SKIP_PATTERNS):
            continue
        if re.match(r"^[A-Za-z][A-Za-z0-9_ ]{2,}$", s):
            return s

    return None


def _event_type_from_suffix(func_name: str) -> str:
    """Map a JS function name suffix to a human-readable event type."""
    mapping = {
        "_OnUp": "OnUp (release)",
        "_OnDown": "OnDown (press)",
        "_OnLoaded": "OnLoaded (screen init)",
        "_OnClick": "OnClick",
        "_OnChange": "OnChange",
        "_OnMouseEnter": "OnMouseEnter",
        "_OnMouseLeave": "OnMouseLeave",
    }
    for suffix, label in mapping.items():
        if suffix in func_name:
            return label
    return "unknown"


def _resolve_tags(
    tag_names: list[str],
    hmi_tags: dict[str, dict],
) -> list[dict]:
    """Resolve a list of tag name strings against the HMITags table."""
    resolved: list[dict] = []
    for tag_name in tag_names:
        tag_detail = hmi_tags.get(tag_name)
        if tag_detail:
            resolved.append(tag_detail)
        else:
            resolved.append(
                {
                    "hmi_tag": tag_name,
                    "plc_tag": "(not found in HMITags.xlsx)",
                }
            )
    return resolved


def _build_screen_elements(elements: list[dict]) -> list[ScreenElement]:
    """Convert raw element dicts into :class:`ScreenElement` models."""
    result: list[ScreenElement] = []
    for elem in elements:
        tag_bindings: list[TagBinding] = []
        for tag in elem.get("hmi_tags", []):
            tag_bindings.append(
                TagBinding(
                    property="process_value",
                    hmi_tag=tag.get("hmi_tag", ""),
                    plc_tag=tag.get("plc_tag", ""),
                    plc_name=tag.get("plc_tag", ""),
                    data_type=tag.get("data_type", ""),
                    connection=tag.get("connection", ""),
                )
            )

        events: list[ElementEvent] = []
        for ev in elem.get("events", []):
            events.append(
                ElementEvent(
                    function=ev.get("function", ""),
                    event_type=ev.get("event_type", ""),
                    plc_tags=ev.get("plc_tags", []),
                    navigates_to=ev.get("navigates_to", []),
                    code=ev.get("code", ""),
                )
            )

        result.append(
            ScreenElement(
                name=elem.get("name", ""),
                type=elem.get("type", ""),
                io_role=elem.get("io_role", ""),
                tag_bindings=tag_bindings,
                events=events,
            )
        )
    return result
